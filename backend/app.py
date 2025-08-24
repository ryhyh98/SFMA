
import sys
import os
import pandas as pd
from flask import Flask, send_from_directory, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference
from io import BytesIO
from fpdf import FPDF
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import logging
import webbrowser
from threading import Timer

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

app = Flask(__name__, static_folder=resource_path(''), static_url_path='')

# --- Logging Setup ---
logging.basicConfig(level=logging.INFO)

# --- Font Setup for Matplotlib and FPDF ---
font_path = resource_path(os.path.join('backend', 'NanumGothic.ttf'))
if os.path.exists(font_path):
    fm.fontManager.addfont(font_path)
    plt.rcParams['font.family'] = 'NanumGothic'
else:
    app.logger.warning(f"Font file not found at {font_path}. Korean text in PDF/chart may not display correctly.")

# Define the path to the CSV file
CSV_FILE_PATH = resource_path('스마트팩토리수준진단_input.csv')

# Define paths for saving results
RESULTS_DIR = resource_path('results')
SURVEYOR_LIST_FILE = resource_path('설문자리스트.xlsx')
os.makedirs(RESULTS_DIR, exist_ok=True)

@app.route('/')
def serve_index():
    return send_from_directory(resource_path(''), 'index.html')

# Load the CSV data once when the app starts
try:
    DIAGNOSIS_QUESTIONS_DF = pd.read_csv(CSV_FILE_PATH)
except FileNotFoundError:
    app.logger.error(f"Error: CSV file not found at {CSV_FILE_PATH}. Please ensure it exists.")
    DIAGNOSIS_QUESTIONS_DF = pd.DataFrame() # Empty DataFrame to prevent errors

# ... (The rest of the app remains the same as before)

@app.route('/api/questions', methods=['GET'])
def get_questions():
    try:
        questions_data = DIAGNOSIS_QUESTIONS_DF.to_dict(orient='records')
        return jsonify(questions_data)
    except Exception as e:
        app.logger.error(f"Error in get_questions: {e}")
        return jsonify({"error": str(e)}), 500

def get_next_serial():
    serial_file = resource_path('serial_number.txt')
    current_serial = 0
    if os.path.exists(serial_file):
        with open(serial_file, 'r') as f:
            try:
                current_serial = int(f.read().strip())
            except ValueError:
                current_serial = 0
    next_serial = current_serial + 1
    with open(serial_file, 'w') as f:
        f.write(str(next_serial))
    return f"SFactory-{str(next_serial).zfill(4)}"

# Helper function for final level name
def get_final_level_name(score):
    LEVEL_DEFINITIONS = [
        {'name': 'Level 0', 'min': 0, 'max': 550},
        {'name': 'Level 1', 'min': 550, 'max': 650},
        {'name': 'Level 2', 'min': 650, 'max': 750},
        {'name': 'Level 3', 'min': 750, 'max': 850},
        {'name': 'Level 4', 'min': 850, 'max': 950},
        {'name': 'Level 5', 'min': 950, 'max': 1001}
    ]
    for level in LEVEL_DEFINITIONS:
        if score < level['max']:
            return level['name']
    return LEVEL_DEFINITIONS[-1]['name']

@app.route('/api/submit_diagnosis', methods=['POST'])
def submit_diagnosis():
    try:
        data = request.json
        app.logger.info(f"Received data: {data}")

        surveyor_name = data.get('surveyorName')
        answers = data.get('answers')

        if not surveyor_name or not answers:
            return jsonify({"error": "Missing surveyorName or answers"}), 400

        total_score = 0
        category_data = {}
        for category in DIAGNOSIS_QUESTIONS_DF['대분류'].unique():
            category_data[category] = {'score': 0, 'level_index_sum': 0, 'question_count': 0}

        for q_no, answer_data in answers.items():
            question_row = DIAGNOSIS_QUESTIONS_DF[DIAGNOSIS_QUESTIONS_DF['No'].astype(str) == q_no]
            if not question_row.empty:
                question_row = question_row.iloc[0]
                category = question_row['대분류']
                score_value = answer_data['value']
                level_index = answer_data['levelIndex']
                total_score += score_value
                category_data[category]['score'] += score_value
                category_data[category]['level_index_sum'] += level_index
                category_data[category]['question_count'] += 1
        
        app.logger.info(f"Calculated category_data: {category_data}")

        category_levels = {cat: round(data['level_index_sum'] / data['question_count'], 2) if data['question_count'] > 0 else 0 for cat, data in category_data.items()}
        category_score_allocations = DIAGNOSIS_QUESTIONS_DF.groupby('대분류')['배점'].first().to_dict()
        serial_number = get_next_serial()
        
        result_for_frontend = {
            "id": serial_number,
            "surveyorName": surveyor_name,
            "date": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "totalScore": total_score,
            "categoryScores": {cat: data['score'] for cat, data in category_data.items()},
            "categoryLevels": category_levels,
            "categoryScoreAllocations": category_score_allocations
        }

        app.logger.info("Preparing to save files.")
        # Save result to a dedicated CSV/Excel file
        result_df = pd.DataFrame([result_for_frontend])
        result_file_path = os.path.join(RESULTS_DIR, f"{serial_number}_{surveyor_name}.csv")
        result_df.to_csv(result_file_path, index=False, encoding='utf-8-sig')
        app.logger.info(f"Saved result CSV to {result_file_path}")

        # Update surveyor list
        surveyor_entry = pd.DataFrame([{
            "일련번호": serial_number,
            "설문자 이름": surveyor_name,
            "날짜": pd.Timestamp.now().strftime('%Y-%m-%d'),
            "총점": total_score,
            "최종 레벨": get_final_level_name(total_score)
        }])

        if os.path.exists(SURVEYOR_LIST_FILE):
            existing_surveyors = pd.read_excel(SURVEYOR_LIST_FILE)
            updated_surveyors = pd.concat([existing_surveyors, surveyor_entry], ignore_index=True)
        else:
            updated_surveyors = surveyor_entry
        
        updated_surveyors.to_excel(SURVEYOR_LIST_FILE, index=False)
        app.logger.info(f"Saved surveyor list to {SURVEYOR_LIST_FILE}")

        return jsonify(result_for_frontend)

    except Exception as e:
        app.logger.error(f"An error occurred in submit_diagnosis: {e}", exc_info=True)
        return jsonify({"error": "An internal server error occurred."}), 500

@app.route('/api/download_excel', methods=['POST'])
def download_excel():
    data = request.json
    wb = Workbook()
    ws = wb.active
    ws.title = "진단결과"

    # --- Styling ---
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # --- Header ---
    headers = ["영역", "배점", "수준", "점수(점)"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        cell.fill = header_fill

    # --- Data ---
    row_idx = 2
    for category in data['categoryLevels']:
        ws.cell(row=row_idx, column=1, value=category).border = border
        ws.cell(row=row_idx, column=2, value=data['categoryScoreAllocations'][category]).border = border
        ws.cell(row=row_idx, column=3, value=data['categoryLevels'][category]).border = border
        ws.cell(row=row_idx, column=4, value=data['categoryScores'][category]).border = border
        row_idx += 1

    # --- Chart ---
    chart = BarChart()
    chart.title = "스마트공장 수준 진단결과(5점 척도 기준)"
    chart.y_axis.title = '수준'
    chart.x_axis.title = '영역'
    
    chart_data = Reference(ws, min_col=3, min_row=2, max_row=row_idx - 1, max_col=3)
    chart_cats = Reference(ws, min_col=1, min_row=2, max_row=row_idx - 1)
    chart.add_data(chart_data, titles_from_data=False)
    chart.set_categories(chart_cats)
    chart.legend = None
    ws.add_chart(chart, "F2")

    # --- Save to BytesIO ---
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    return send_file(
        excel_io,
        as_attachment=True,
        download_name=f"{data['id']}_진단결과.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/download_pdf', methods=['POST'])
def download_pdf():
    data = request.json
    
    # --- Create Chart Image ---
    plt.figure(figsize=(8, 4))
    plt.bar(data['categoryLevels'].keys(), data['categoryLevels'].values(), color='skyblue')
    plt.title('스마트공장 수준 진단결과(5점 척도 기준)')
    plt.ylabel('수준')
    plt.xlabel('영역')
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    
    img_io = BytesIO()
    plt.savefig(img_io, format='png')
    img_io.seek(0)
    plt.close()

    # --- Create PDF ---
    pdf = FPDF()
    pdf.add_page()
    if os.path.exists(font_path):
        pdf.add_font('NanumGothic', '', font_path, uni=True)
        pdf.set_font('NanumGothic', size=16)
    else: # Fallback font
        pdf.set_font('Arial', size=16)

    pdf.cell(0, 10, '스마트공장 수준 진단 결과', 0, 1, 'C')
    pdf.ln(10)

    # --- Table ---
    if os.path.exists(font_path):
        pdf.set_font('NanumGothic', size=12)
    else:
        pdf.set_font('Arial', size=12)

    col_widths = [60, 30, 30, 30]
    headers = ["영역", "배점", "수준", "점수(점)"]
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
    pdf.ln()

    for category in data['categoryLevels']:
        pdf.cell(col_widths[0], 10, category, 1, 0)
        pdf.cell(col_widths[1], 10, str(data['categoryScoreAllocations'][category]), 1, 0, 'C')
        pdf.cell(col_widths[2], 10, str(data['categoryLevels'][category]), 1, 0, 'C')
        pdf.cell(col_widths[3], 10, str(data['categoryScores'][category]), 1, 1, 'C')

    pdf.ln(10)
    pdf.image(img_io, x=10, y=pdf.get_y(), w=190)

    # --- Save to BytesIO ---
    pdf_io = BytesIO()
    pdf.output(pdf_io)
    pdf_io.seek(0)

    return send_file(
        pdf_io,
        as_attachment=True,
        download_name=f"{data['id']}_진단결과.pdf",
        mimetype='application/pdf'
    )

def open_browser():
    webbrowser.open_new("http://127.0.0.1:5000/")

if __name__ == '__main__':
    Timer(1, open_browser).start()
    app.run(debug=False, port=5000)
